from __future__ import annotations

import math
import pathlib
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

from azure_estate.naming import run_stamp

# Excel refuses any cell longer than this; openpyxl would silently truncate it.
_MAX_CELL = 32767
_ELLIPSIS = "… (truncado)"

# Caracteres de controle que o Excel recusa; openpyxl levanta
# IllegalCharacterError e derruba a geração inteira por causa de uma célula.
_ILEGAIS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Quantas linhas bastam para estimar a largura de uma coluna.  Varrer as
# 39 mil linhas de uma aba grande custa tempo e não muda o resultado, já que a
# largura é limitada a 80 caracteres.
_AMOSTRA_LARGURA = 200
_LARGURA_MAX = 80


def _limpa(valor):
    """Devolve *valor* aceitável para uma célula do Excel."""
    if valor is None:
        return None
    if isinstance(valor, float) and math.isnan(valor):
        return None
    if isinstance(valor, (int, float, bool)):
        return valor
    texto = valor if isinstance(valor, str) else str(valor)
    if _ILEGAIS.search(texto):
        texto = _ILEGAIS.sub("", texto)
    if len(texto) > _MAX_CELL:
        texto = texto[: _MAX_CELL - len(_ELLIPSIS)] + _ELLIPSIS
    return texto


def _larguras(df: pd.DataFrame) -> list[float]:
    """Largura de cada coluna, estimada a partir do cabeçalho e de uma amostra."""
    amostra = df.head(_AMOSTRA_LARGURA)
    larguras = []
    for coluna in df.columns:
        maior = len(str(coluna))
        for valor in amostra[coluna]:
            if valor is None:
                continue
            tamanho = len(str(valor))
            if tamanho > maior:
                maior = tamanho
                if maior >= _LARGURA_MAX:
                    break
        larguras.append(min(maior + 4, _LARGURA_MAX))
    return larguras


def _escreve_aba(worksheet, df: pd.DataFrame) -> None:
    """Escreve *df* numa planilha em modo write-only, linha a linha.

    O modo write-only não guarda as células escritas, então o pico de memória
    passa a ser o de uma linha e não o do livro inteiro.
    """
    for indice, largura in enumerate(_larguras(df), start=1):
        worksheet.column_dimensions[get_column_letter(indice)].width = largura

    worksheet.append([str(c) for c in df.columns])
    for linha in df.itertuples(index=False, name=None):
        worksheet.append([_limpa(valor) for valor in linha])


def _sanitiza(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica ``_limpa`` às colunas de texto de *df*.

    Usado no caminho com gráfico, que precisa do modo normal do openpyxl e por
    isso não passa por ``_escreve_aba``.
    """
    out = df.copy()
    for coluna in out.columns:
        if pd.api.types.is_numeric_dtype(out[coluna]) or pd.api.types.is_bool_dtype(
            out[coluna]
        ):
            continue
        out[coluna] = out[coluna].map(_limpa)
    return out


def _clip_cells(df: pd.DataFrame) -> pd.DataFrame:
    """Shorten oversized text cells, flagging that content was cut.

    Mantido para quem já dependia da função; o caminho de escrita agora corta
    célula a célula em ``_limpa``, sem copiar o DataFrame.
    """
    out = df.copy()
    for column in out.columns:
        values = out[column]
        if pd.api.types.is_numeric_dtype(values) or pd.api.types.is_bool_dtype(values):
            continue
        text = values.astype(str)
        if not (text.str.len() > _MAX_CELL).any():
            continue
        out[column] = text.mask(
            text.str.len() > _MAX_CELL,
            text.str.slice(0, _MAX_CELL - len(_ELLIPSIS)) + _ELLIPSIS,
        )
    return out


class ExcelExporter:
    """Saves one or more DataFrames as sheets in an Excel workbook."""

    def __init__(self, output_dir: str = "output") -> None:
        self._output_dir = pathlib.Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

    def save(
        self,
        df: pd.DataFrame,
        sheet_name: str = "Sheet1",
        filename: str | None = None,
    ) -> pathlib.Path:
        """Write *df* to an Excel file and return the resolved path.

        Parameters
        ----------
        df:
            DataFrame to export.
        sheet_name:
            Name of the Excel worksheet.
        filename:
            Base filename (without directory).  Defaults to
            ``<sheet_name>_DD_MM_YYYY_HH_MM_SS.xlsx``.
        """
        if filename is None:
            filename = f"{sheet_name}_{run_stamp()}.xlsx"

        path = self._output_dir / filename

        workbook = Workbook(write_only=True)
        try:
            _escreve_aba(workbook.create_sheet(title=sheet_name[:31]), df)
            workbook.save(path)
        finally:
            workbook.close()

        return path

    @staticmethod
    def _auto_fit(worksheet) -> None:
        """Auto-fit column widths for a worksheet (max 80 chars)."""
        for col_cells in worksheet.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            worksheet.column_dimensions[col_cells[0].column_letter].width = (
                min(max_len + 4, _LARGURA_MAX)
            )

    def save_multi_sheet(
        self,
        sheets: list[tuple[str, "pd.DataFrame"]],
        filename: str,
    ) -> pathlib.Path:
        """Write multiple DataFrames as separate sheets in one Excel workbook.

        Parameters
        ----------
        sheets:
            Ordered list of (sheet_name, DataFrame). Empty DataFrames are skipped.
        filename:
            Base filename (without directory).
        """
        path = self._output_dir / filename

        workbook = Workbook(write_only=True)
        try:
            for sheet_name, df in sheets:
                if df is None or df.empty:
                    continue
                _escreve_aba(workbook.create_sheet(title=sheet_name[:31]), df)
            workbook.save(path)
        finally:
            workbook.close()

        return path

    def save_with_pie_chart(
        self,
        df: "pd.DataFrame",
        label_col: str,
        value_col: str,
        sheet_name: str = "Sheet1",
        chart_sheet_name: str = "Gráfico",
        chart_title: str = "Distribuição",
        top_n: int = 15,
        filename: str | None = None,
    ) -> pathlib.Path:
        """Write *df* to Excel with a pie chart on a second sheet.

        Parameters
        ----------
        df:
            Full DataFrame (already sorted as desired).
        label_col:
            Column to use as pie slice labels.
        value_col:
            Column to use as pie slice values.
        top_n:
            Keep the top N rows in the chart; remaining rows are grouped as
            'Outros' to keep the chart readable.
        """
        if filename is None:
            filename = f"{sheet_name}_{run_stamp()}.xlsx"

        path = self._output_dir / filename

        # --- Build chart DataFrame (top N + Outros) ---
        chart_df = df[[label_col, value_col]].copy()
        if len(chart_df) > top_n:
            top = chart_df.head(top_n).copy()
            outros_total = chart_df.iloc[top_n:][value_col].sum()
            outros_row = pd.DataFrame([{label_col: "Outros", value_col: outros_total}])
            chart_df = pd.concat([top, outros_row], ignore_index=True)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # Sheet 1: full data
            _sanitiza(df).to_excel(writer, index=False, sheet_name=sheet_name[:31])
            self._auto_fit(writer.sheets[sheet_name[:31]])

            # Sheet 2: chart source data (hidden helper sheet) + chart
            chart_df.to_excel(writer, index=False, sheet_name="_chart_data")
            ws_chart_data = writer.sheets["_chart_data"]
            ws_chart_data.sheet_state = "hidden"

            # Build the pie chart referencing the hidden sheet
            n_rows = len(chart_df) + 1  # +1 for header
            labels = Reference(ws_chart_data, min_col=1, min_row=2, max_row=n_rows)
            data = Reference(ws_chart_data, min_col=2, min_row=1, max_row=n_rows)

            pie = PieChart()
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = chart_title
            pie.style = 10
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.dataLabels.showCatName = False
            pie.dataLabels.showVal = False
            pie.dataLabels.showSerName = False
            pie.width = 22
            pie.height = 16

            # Create chart sheet
            wb = writer.book
            ws_chart = wb.create_sheet(title=chart_sheet_name[:31])
            ws_chart.add_chart(pie, "B2")

        return path
