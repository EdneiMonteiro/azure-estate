"""Testa o exportador Excel: conteudo correto e pico de memoria baixo.

Roda sem Azure. Uso: python testa-excel.py
"""

from __future__ import annotations

import shutil
import sys
import tempfile
import tracemalloc
from pathlib import Path

import openpyxl
import pandas as pd

from azure_estate.exporters.excel import ExcelExporter, _limpa

falhas: list[str] = []
total = 0


def checa(rotulo: str, obtido, esperado) -> None:
    global total
    total += 1
    if obtido == esperado:
        print(f"  ok   {rotulo}: {obtido!r}")
    else:
        print(f"  FALHA {rotulo}: obtido {obtido!r}, esperado {esperado!r}")
        falhas.append(rotulo)


def checa_true(rotulo: str, condicao: bool, detalhe: str = "") -> None:
    global total
    total += 1
    if condicao:
        print(f"  ok   {rotulo}{(' ' + detalhe) if detalhe else ''}")
    else:
        print(f"  FALHA {rotulo} {detalhe}")
        falhas.append(rotulo)


def le(caminho: Path) -> dict[str, list[tuple]]:
    wb = openpyxl.load_workbook(caminho, read_only=True)
    dados = {n: list(wb[n].iter_rows(values_only=True)) for n in wb.sheetnames}
    wb.close()
    return dados


saida = Path(tempfile.mkdtemp(prefix="azestate-xlsx-"))
exportador = ExcelExporter(output_dir=str(saida))

try:
    print("1. Conteudo preservado: cabecalho, linhas, tipos")
    df = pd.DataFrame(
        {
            "Nome": ["vm-a", "vm-b"],
            "vCPUs": [2, 16],
            "Ativo": [True, False],
            "Zonas": ["1; 2", ""],
        }
    )
    caminho = exportador.save(df, sheet_name="VMs", filename="um.xlsx")
    dados = le(caminho)
    checa("uma aba", list(dados), ["VMs"])
    checa("cabecalho", dados["VMs"][0], ("Nome", "vCPUs", "Ativo", "Zonas"))
    checa("linha 1", dados["VMs"][1], ("vm-a", 2, True, "1; 2"))
    checa("numero segue numero", isinstance(dados["VMs"][1][1], int), True)
    checa("booleano segue booleano", isinstance(dados["VMs"][1][2], bool), True)

    print("\n2. Multiplas abas: vazias sao puladas, nome cortado em 31")
    abas = [
        ("Primeira", pd.DataFrame({"a": [1]})),
        ("Vazia", pd.DataFrame()),
        ("Nula", None),
        ("N" * 40, pd.DataFrame({"b": [2]})),
    ]
    caminho = exportador.save_multi_sheet(abas, "multi.xlsx")
    dados = le(caminho)
    checa("abas escritas", list(dados), ["Primeira", "N" * 31])
    checa("valor da segunda", dados["N" * 31][1], (2,))

    print("\n3. Caractere ilegal nao derruba a geracao (antes: IllegalCharacterError)")
    sujo = pd.DataFrame({"txt": ["ok\x00\x07mesmo assim", "limpo"]})
    caminho = exportador.save_multi_sheet([("Sujo", sujo)], "sujo.xlsx")
    dados = le(caminho)
    checa("controle removido", dados["Sujo"][1][0], "okmesmo assim")
    checa("linha intacta", dados["Sujo"][2][0], "limpo")

    print("\n4. Celula acima do limite do Excel e cortada com marca")
    gigante = pd.DataFrame({"txt": ["x" * 40000]})
    caminho = exportador.save_multi_sheet([("Gigante", gigante)], "gigante.xlsx")
    valor = le(caminho)["Gigante"][1][0]
    checa("dentro do limite", len(valor) <= 32767, True)
    checa("marcado como truncado", valor.endswith("(truncado)"), True)

    print("\n5. Vazio/NaN vira celula vazia, nao a string 'nan'")
    com_nan = pd.DataFrame({"a": [1.0, float("nan")], "b": ["x", None]})
    caminho = exportador.save_multi_sheet([("Nan", com_nan)], "nan.xlsx")
    aba = le(caminho)["Nan"]
    checa("linha vazia nao some", len(aba), 3)
    linha = tuple(aba[2]) + (None,) * (2 - len(aba[2]))
    checa("nan virou vazio", linha[0], None)
    checa("none virou vazio", linha[1], None)
    lidas = pd.read_excel(caminho, sheet_name="Nan")
    # Linha inteiramente vazia no fim some para o pandas — comportamento
    # identico ao antigo pd.to_excel, medido: e do formato, nao do exportador.
    checa("mesma leitura do to_excel antigo", len(lidas), 1)
    checa("sem a string nan", "nan" in lidas.astype(str).values, False)

    print("\n6. Pico de memoria escala com a linha, nao com o livro")
    grande = pd.DataFrame(
        {f"col{i}": [f"valor-{i}-{j}" for j in range(4000)] for i in range(12)}
    )
    muitas = [(f"Aba{i}", grande) for i in range(8)]  # 32 mil linhas x 12 colunas

    tracemalloc.start()
    base = tracemalloc.get_traced_memory()[0]
    caminho = exportador.save_multi_sheet(muitas, "grande.xlsx")
    pico = (tracemalloc.get_traced_memory()[1] - base) / 1024 / 1024
    tracemalloc.stop()

    linhas = sum(len(d) for _, d in muitas)
    print(f"       {linhas} linhas em {len(muitas)} abas -> pico {pico:.1f} MB")
    checa_true(
        "pico abaixo de 40 MB",
        pico < 40,
        f"({pico:.1f} MB)",
    )
    dados = le(caminho)
    checa("todas as abas gravadas", len(dados), 8)
    checa("linhas por aba", len(dados["Aba0"]) - 1, 4000)
    checa("ultima linha correta", dados["Aba7"][-1][0], "valor-0-3999")

    print("\n7. Larguras de coluna definidas e limitadas a 80")
    wb = openpyxl.load_workbook(caminho)
    larguras = [
        d.width for d in wb["Aba0"].column_dimensions.values() if d.width is not None
    ]
    wb.close()
    checa("uma largura por coluna", len(larguras), 12)
    checa_true("nenhuma acima de 80", all(w <= 80 for w in larguras), str(max(larguras)))

    print("\n8. _limpa nao altera o que ja esta correto")
    checa("texto comum", _limpa("Standard_LRS"), "Standard_LRS")
    checa("inteiro", _limpa(7), 7)
    checa("quebra de linha preservada", _limpa("a\nb"), "a\nb")
    checa("tabulacao preservada", _limpa("a\tb"), "a\tb")

finally:
    shutil.rmtree(saida, ignore_errors=True)

print("\n" + "=" * 60)
if falhas:
    print(f"{len(falhas)} de {total} casos falharam: {', '.join(falhas)}")
    sys.exit(1)
print(f"{total}/{total} casos passaram.")
